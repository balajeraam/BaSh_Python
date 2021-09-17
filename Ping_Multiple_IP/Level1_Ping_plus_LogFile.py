#import relevant modules
import openpyxl
import time
import datetime
import subprocess
import platform

# Requesting user to provide the path where XLSX file containing server ip list is stored
path = input("Enter path where .XLSX containing server list is stored  >> ")

# In Windows separator used is "\". We need to modify it to separator "/" that python understands
path = path.replace(os.sep,"/")

# Loading active worksheet in the .xlsx file
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

# In our excel file first row is header. So range starts from 2.
for row in range (2,sheet_obj.max_row+1):
    
    ip_address_cell = sheet_obj.cell(row,1)
    
    # Building the command that will be used to ping the ipaddress
    if platform.system().lower()=="windows" :
        parameter ='-n'
    else :
        parameter ='-c'
    command = ['ping',parameter,'3', ip_address_cell.value]
   
    run_time = datetime.datetime.now()
    timestamp_cell = sheet_obj.cell(row,4)
    timestamp_cell.value = run_time
    
    # Ping and obtain system code
    outcome = subprocess.call((command),stdout=subprocess.DEVNULL)
    ip_status_cell = sheet_obj.cell(row,5)
    ip_status_cell.value = outcome
    
    # Translating the status code into comments that can also help a non-technical personal
    comment_cell = sheet_obj.cell(row,6)
    if outcome == 0 :
        comment_cell.value = "Reachable"
    else :
        comment_cell.value = "Not Reachable"
    
    wb_obj.save(path)
    # Small wait time provided after update for ipaddress is complete
    time.sleep(2)
